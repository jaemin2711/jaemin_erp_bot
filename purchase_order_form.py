from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


PURCHASE_FILE = Path("data/purchase_request.xlsx")
ORDER_DIR = Path("data/purchase_orders")


BUYER_INFO = {
    "업체명": "",
    "대표자": "",
    "사업자번호": "",
    "주소": "",
    "담당자": "",
    "TEL": "",
    "FAX": "",
    "E-Mail": "",
}


SUPPLIER_INFO = {
    "업체명": "",
    "대표자": "",
    "사업자번호": "",
    "주소": "",
    "담당자": "",
    "TEL": "",
    "FAX": "",
    "E-Mail": "",
}


def fmt_num(value):
    try:
        value = float(value)
        if value.is_integer():
            return f"{value:,.0f}"
        return f"{value:,.3f}".rstrip("0").rstrip(".")
    except Exception:
        return str(value)


def to_float(value):
    try:
        if value is None:
            return 0
        text = str(value).replace(",", "").strip()
        if text == "" or text.lower() == "nan":
            return 0
        return float(text)
    except Exception:
        return 0


def make_border():
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def set_cell(ws, cell, value=None, bold=False, size=10, align="center", fill=None):
    ws[cell] = value
    ws[cell].font = Font(name="맑은 고딕", size=size, bold=bold)
    ws[cell].alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    ws[cell].border = make_border()

    if fill:
        ws[cell].fill = PatternFill("solid", fgColor=fill)


def merge_set(ws, cell_range, value=None, bold=False, size=10, align="center", fill=None):
    ws.merge_cells(cell_range)
    start_cell = cell_range.split(":")[0]
    ws[start_cell] = value
    ws[start_cell].font = Font(name="맑은 고딕", size=size, bold=bold)
    ws[start_cell].alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    for row in ws[cell_range]:
        for cell in row:
            cell.border = make_border()
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)


def format_date(value):
    if value is None:
        return ""

    try:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass

    return str(value)


def load_purchase_requests():
    if not PURCHASE_FILE.exists():
        raise FileNotFoundError(f"발주요청 파일이 없습니다: {PURCHASE_FILE}")

    df = pd.read_excel(PURCHASE_FILE)

    if df.empty:
        raise ValueError("발주요청 파일에 데이터가 없습니다.")

    return df


def pick_purchase_rows(purchase_no=None):
    df = load_purchase_requests()

    if "발주요청번호" not in df.columns:
        raise ValueError("purchase_request.xlsx에 '발주요청번호' 컬럼이 없습니다.")

    if purchase_no:
        target = str(purchase_no).strip()
        rows = df[df["발주요청번호"].astype(str).str.strip() == target].copy()

        if rows.empty:
            raise ValueError(f"발주요청번호를 찾지 못했습니다: {purchase_no}")

        return rows

    # 발주번호를 지정하지 않으면 가장 최근 발주요청번호 사용
    if "발주요청일시" in df.columns:
        df["_정렬일시"] = pd.to_datetime(df["발주요청일시"], errors="coerce")
        df = df.sort_values("_정렬일시", ascending=False)

    latest_no = str(df.iloc[0]["발주요청번호"]).strip()
    rows = df[df["발주요청번호"].astype(str).str.strip() == latest_no].copy()

    return rows


def setup_page(ws):
    ws.title = "구매발주서"

    # 열 너비
    widths = {
        "A": 4,
        "B": 9,
        "C": 13,
        "D": 13,
        "E": 12,
        "F": 12,
        "G": 10,
        "H": 12,
        "I": 10,
        "J": 12,
        "K": 12,
        "L": 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 행 높이
    for row in range(1, 45):
        ws.row_dimensions[row].height = 22

    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 30

    # 인쇄 설정
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_view.showGridLines = False


def draw_title_and_approval(ws):
    # 제목
    ws.merge_cells("B2:F3")
    ws["B2"] = "구  매  발  주  서"
    ws["B2"].font = Font(name="맑은 고딕", size=20, bold=True)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    # 제목 밑줄 느낌
    thin = Side(style="medium", color="000000")
    for col in range(2, 7):
        ws.cell(row=3, column=col).border = Border(bottom=thin)

    # 결재란
    gray = "D9D9D9"
    merge_set(ws, "I2:J2", "작성", bold=True, fill=gray)
    merge_set(ws, "K2:L2", "검토", bold=True, fill=gray)
    merge_set(ws, "M2:N2", "승인", bold=True, fill=gray)

    merge_set(ws, "I3:J5", "")
    merge_set(ws, "K3:L5", "")
    merge_set(ws, "M3:N5", "")


def draw_order_header(ws, rows):
    gray = "D9D9D9"

    first = rows.iloc[0]

    purchase_no = str(first.get("발주요청번호", "")).strip()
    request_dt = first.get("발주요청일시", datetime.now())
    order_date = format_date(request_dt)

    merge_set(ws, "A7:B7", "발주번호", bold=True, fill=gray)
    merge_set(ws, "C7:E7", purchase_no)
    merge_set(ws, "F7:G7", "발주일", bold=True, fill=gray)
    merge_set(ws, "H7:J7", order_date)


def draw_company_boxes(ws):
    gray = "D9D9D9"

    # 발주처 / 공급처 타이틀
    merge_set(ws, "A8:F8", "발 주 처", bold=True, fill=gray)
    merge_set(ws, "H8:N8", "공 급 처", bold=True, fill=gray)

    # 발주처 기본정보
    merge_set(ws, "A9:B9", "업체명", bold=True, fill=gray)
    merge_set(ws, "C9:E9", BUYER_INFO["업체명"], align="left")
    merge_set(ws, "F9:F9", "대표자", bold=True, fill=gray)
    merge_set(ws, "G9:G9", BUYER_INFO["대표자"], align="left")

    merge_set(ws, "A10:B10", "사업자번호", bold=True, fill=gray)
    merge_set(ws, "C10:G10", BUYER_INFO["사업자번호"], align="left")

    merge_set(ws, "A11:B11", "주소", bold=True, fill=gray)
    merge_set(ws, "C11:G11", BUYER_INFO["주소"], align="left")

    # 공급처 기본정보
    merge_set(ws, "H9:I9", "업체명", bold=True, fill=gray)
    merge_set(ws, "J9:L9", SUPPLIER_INFO["업체명"], align="left")
    merge_set(ws, "M9:M9", "대표자", bold=True, fill=gray)
    merge_set(ws, "N9:N9", SUPPLIER_INFO["대표자"], align="left")

    merge_set(ws, "H10:I10", "사업자번호", bold=True, fill=gray)
    merge_set(ws, "J10:N10", SUPPLIER_INFO["사업자번호"], align="left")

    merge_set(ws, "H11:I11", "주소", bold=True, fill=gray)
    merge_set(ws, "J11:N11", SUPPLIER_INFO["주소"], align="left")

    # 담당자 영역
    merge_set(ws, "A12:B15", "발 주\n담당자", bold=True, fill=gray)
    merge_set(ws, "C12:D12", "NAME", bold=True)
    merge_set(ws, "E12:G12", BUYER_INFO["담당자"], align="left")
    merge_set(ws, "C13:D13", "TEL", bold=True)
    merge_set(ws, "E13:G13", BUYER_INFO["TEL"], align="left")
    merge_set(ws, "C14:D14", "FAX", bold=True)
    merge_set(ws, "E14:G14", BUYER_INFO["FAX"], align="left")
    merge_set(ws, "C15:D15", "E-Mail", bold=True)
    merge_set(ws, "E15:G15", BUYER_INFO["E-Mail"], align="left")

    merge_set(ws, "H12:I15", "공 급\n담당자", bold=True, fill=gray)
    merge_set(ws, "J12:K12", "NAME", bold=True)
    merge_set(ws, "L12:N12", SUPPLIER_INFO["담당자"], align="left")
    merge_set(ws, "J13:K13", "TEL", bold=True)
    merge_set(ws, "L13:N13", SUPPLIER_INFO["TEL"], align="left")
    merge_set(ws, "J14:K14", "FAX", bold=True)
    merge_set(ws, "L14:N14", SUPPLIER_INFO["FAX"], align="left")
    merge_set(ws, "J15:K15", "E-Mail", bold=True)
    merge_set(ws, "L15:N15", SUPPLIER_INFO["E-Mail"], align="left")

    # 특이사항
    merge_set(ws, "A16:B18", "특이사항", bold=True, fill=gray)
    merge_set(ws, "C16:N18", "", align="left")


def draw_item_table(ws, rows):
    gray = "D9D9D9"

    header_row = 19

    headers = [
        ("A", "A", "순\n번"),
        ("B", "C", "품의번호"),
        ("D", "F", "품명\n품번"),
        ("G", "H", "규격\n단위"),
        ("I", "I", "요청수량\n납기일"),
        ("J", "J", "부가세\n포함"),
        ("K", "L", "단가\n금액 (KRW)"),
        ("M", "N", "제조사\n창고"),
    ]

    for start_col, end_col, title in headers:
        cell_range = f"{start_col}{header_row}:{end_col}{header_row + 1}"
        merge_set(ws, cell_range, title, bold=True, fill=gray)

    start_row = 21

    for idx, (_, item) in enumerate(rows.iterrows(), start=1):
        r1 = start_row + (idx - 1) * 2
        r2 = r1 + 1

        purchase_no = str(item.get("발주요청번호", ""))
        material_code = str(item.get("자재코드", ""))
        material_name = str(item.get("자재명", ""))

        spec = str(item.get("규격", "")) if "규격" in rows.columns else ""
        unit = str(item.get("단위", ""))

        qty = to_float(item.get("발주요청수량", item.get("부족수량", 0)))
        due_date = format_date(item.get("생산일", ""))

        unit_price = to_float(item.get("단가", 0)) if "단가" in rows.columns else 0
        amount = qty * unit_price if unit_price else 0

        maker = str(item.get("제조사", "")) if "제조사" in rows.columns else ""
        warehouse = str(item.get("창고", "원/부자재창고")) if "창고" in rows.columns else "원/부자재창고"

        merge_set(ws, f"A{r1}:A{r2}", idx)
        merge_set(ws, f"B{r1}:C{r2}", purchase_no, size=8)
        merge_set(ws, f"D{r1}:F{r1}", material_name, align="left")
        merge_set(ws, f"D{r2}:F{r2}", material_code, align="left")
        merge_set(ws, f"G{r1}:H{r1}", spec, align="left")
        merge_set(ws, f"G{r2}:H{r2}", unit)
        merge_set(ws, f"I{r1}:I{r1}", qty)
        merge_set(ws, f"I{r2}:I{r2}", due_date)
        merge_set(ws, f"J{r1}:J{r2}", "")
        merge_set(ws, f"K{r1}:L{r1}", unit_price if unit_price else "")
        merge_set(ws, f"K{r2}:L{r2}", amount if amount else "")
        merge_set(ws, f"M{r1}:N{r1}", maker)
        merge_set(ws, f"M{r2}:N{r2}", warehouse)

        ws[f"I{r1}"].number_format = "#,##0.###"
        ws[f"K{r1}"].number_format = "#,##0"
        ws[f"K{r2}"].number_format = "#,##0"

    # 빈 줄 추가
    used_rows = start_row + len(rows) * 2
    for r in range(used_rows, used_rows + 8):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.border = make_border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="맑은 고딕", size=9)


def create_purchase_order_excel(purchase_no=None):
    rows = pick_purchase_rows(purchase_no=purchase_no)

    first = rows.iloc[0]
    purchase_no = str(first.get("발주요청번호", "")).strip()

    wb = Workbook()
    ws = wb.active

    setup_page(ws)
    draw_title_and_approval(ws)
    draw_order_header(ws, rows)
    draw_company_boxes(ws)
    draw_item_table(ws, rows)

    ORDER_DIR.mkdir(parents=True, exist_ok=True)

    today = datetime.now().strftime("%Y%m%d")
    out_file = ORDER_DIR / f"구매발주서_{purchase_no}_{today}.xlsx"

    wb.save(out_file)

    return out_file

def is_purchase_order_form_question(question: str):
    """
    구매발주서 생성 명령 감지.
    예:
    - 발주서 만들어줘
    - 구매발주서 만들어줘
    - 발주서 생성
    - 발주서 출력
    - 발주서 엑셀 만들어줘
    """
    q = str(question).replace(" ", "")

    if "발주서" not in q and "구매발주서" not in q:
        return False

    keywords = [
        "만들",
        "생성",
        "출력",
        "엑셀",
        "작성",
        "뽑아",
    ]

    return any(word in q for word in keywords)

if __name__ == "__main__":
    purchase_no = input("발주요청번호를 입력하세요. 비우면 최신 발주요청으로 생성합니다: ").strip()

    if not purchase_no:
        purchase_no = None

    out_file = create_purchase_order_excel(purchase_no=purchase_no)

    print()
    print("구매발주서 생성 완료")
    print(f"저장 위치: {out_file}")